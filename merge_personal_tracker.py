"""Merge Brandon's personal Master Tracker into the dashboard workbook.

Imports:
- His current portfolio (Cost Analysis sheet → Inventory sheet)
- His Trade Log (Trade Log sheet → new Personal Trades sheet)

Both stay separate from the public Watchlist so Pro subscribers don't see
his private positions.
"""

from datetime import datetime, date
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re

TRACKER = "/sessions/jolly-charming-brahmagupta/mnt/uploads/Pokemon_Master_Tracker_UPDATED_05-07-26.xlsx"
DEST = "Pokemon_Card_Investment_Workbook.xlsx"


def parse_date(v):
    if isinstance(v, (datetime, date)):
        return v
    if isinstance(v, str):
        # "Apr 2026" → 2026-04-01
        m = re.match(r"^(\w{3})\s*(\d{4})$", v.strip())
        if m:
            try:
                return datetime.strptime(f"{m.group(1)} 1 {m.group(2)}", "%b %d %Y")
            except Exception:
                pass
    return datetime(2026, 5, 7)


def main():
    src = load_workbook(TRACKER, data_only=True)
    dst = load_workbook(DEST)

    # ============== 1. Personal Inventory → Inventory sheet ==============
    cost = src["💸 Cost Analysis"]
    inv = dst["Inventory"]

    # Find next row in Inventory (existing data has sample positions)
    # Headers are at row 4, data starts at row 5
    next_inv_row = 5
    while inv.cell(next_inv_row, 1).value:
        next_inv_row += 1

    print(f"Inventory currently has {next_inv_row - 5} rows; appending personal positions...")

    added_inv = 0
    seen = set()
    for r in range(7, cost.max_row + 1):
        dt = cost.cell(r, 1).value
        category = cost.cell(r, 2).value
        item = cost.cell(r, 3).value
        platform = cost.cell(r, 4).value
        spent = cost.cell(r, 5).value or 0
        revenue = cost.cell(r, 6).value or 0
        curr_val = cost.cell(r, 7).value or 0
        in_port = cost.cell(r, 8).value
        graded = cost.cell(r, 9).value
        notes = cost.cell(r, 10).value or ""

        if not item:
            continue
        if str(in_port).upper() not in ("Y", "YES", "TRUE"):
            continue
        key = (str(item)[:40], spent, curr_val)
        if key in seen:
            continue
        seen.add(key)

        rr = next_inv_row
        item_id = f"P-{added_inv+1:03d}"
        cost_unit = spent if spent else curr_val
        # Map columns: 1=Item ID, 2=CardID, 3=Match, 4=Description,
        # 5=Acquired, 6=Qty, 7=Cost/Unit, 8=Total Cost,
        # 9=Current/Unit, 10=Current Total, 11=Current Value,
        # 12=P&L $, 13=P&L %, 14=Days Held, 15=...., 18=Buyer
        inv.cell(rr, 1, item_id)
        inv.cell(rr, 2, "PERSONAL")
        inv.cell(rr, 4, str(item)[:80])
        inv.cell(rr, 5, parse_date(dt))
        inv.cell(rr, 6, 1)
        inv.cell(rr, 7, float(cost_unit))
        inv.cell(rr, 8, float(cost_unit))
        inv.cell(rr, 9, float(curr_val))
        inv.cell(rr, 11, float(curr_val))
        inv.cell(rr, 12, float(curr_val) - float(cost_unit))
        inv.cell(rr, 13, (float(curr_val) - float(cost_unit)) / float(cost_unit) if cost_unit else 0)
        inv.cell(rr, 18, "Brandon")
        # Notes column not always present — write to last column
        inv.cell(rr, 16, f"{category} · {platform or ''} · {notes}".strip(" ·"))
        next_inv_row += 1
        added_inv += 1

    print(f"  ✓ Added {added_inv} personal positions to Inventory")

    # ============== 2. Trade Log → new "Personal Trades" sheet ==============
    if "Personal Trades" in dst.sheetnames:
        del dst["Personal Trades"]
    trades = dst.create_sheet("Personal Trades")
    bold = Font(bold=True, color="FFFFFF")
    gold = PatternFill(start_color="C9A73C", end_color="C9A73C", fill_type="solid")

    trades.cell(1, 1, "🔄 PERSONAL TRADE LOG").font = Font(bold=True, size=14)
    headers = ["Date", "Type", "Cards In", "Cards Out / Cash Paid",
               "Value In $", "Value Out $", "Net Gain $",
               "Platform / Partner", "Condition", "Grade Result", "Strategy / Notes"]
    for i, h in enumerate(headers, 1):
        c = trades.cell(3, i, h)
        c.font = bold
        c.fill = gold
        c.alignment = Alignment(horizontal="center")
        trades.column_dimensions[get_column_letter(i)].width = 18

    src_trades = src["🔄 Trade Log"]
    out_row = 4
    added_trades = 0
    for r in range(4, src_trades.max_row + 1):
        type_ = src_trades.cell(r, 2).value
        cards_in = src_trades.cell(r, 3).value
        if not (type_ or cards_in):
            continue
        dt = src_trades.cell(r, 1).value
        trades.cell(out_row, 1, parse_date(dt) if dt else None)
        trades.cell(out_row, 2, type_)
        trades.cell(out_row, 3, cards_in)
        trades.cell(out_row, 4, src_trades.cell(r, 4).value)
        trades.cell(out_row, 5, src_trades.cell(r, 5).value)
        trades.cell(out_row, 6, src_trades.cell(r, 6).value)
        trades.cell(out_row, 7, src_trades.cell(r, 7).value)
        trades.cell(out_row, 8, src_trades.cell(r, 8).value)
        trades.cell(out_row, 9, src_trades.cell(r, 9).value)
        trades.cell(out_row, 10, src_trades.cell(r, 10).value)
        trades.cell(out_row, 11, src_trades.cell(r, 11).value)
        out_row += 1
        added_trades += 1

    print(f"  ✓ Added {added_trades} trades to Personal Trades sheet")

    # ============== 3. Hunt List → new "Hunt List" sheet ==============
    # Take the user's Chase Cards as the hunt list (cards they want but don't own)
    if "Hunt List" in dst.sheetnames:
        del dst["Hunt List"]
    hunt = dst.create_sheet("Hunt List")
    hunt.cell(1, 1, "🎯 HUNT LIST — Cards we want to acquire").font = Font(bold=True, size=14)
    hunt_headers = ["Card Name", "Set", "Year", "Card #", "Era / Type",
                    "Raw NM ($)", "PSA 10 ($)", "PSA 9 ($)", "Notes"]
    for i, h in enumerate(hunt_headers, 1):
        c = hunt.cell(3, i, h)
        c.font = bold
        c.fill = gold
        c.alignment = Alignment(horizontal="center")
        hunt.column_dimensions[get_column_letter(i)].width = 22

    chase = src["🔥 Chase Cards"]
    out_row = 4
    added_chase = 0
    for r in range(6, chase.max_row + 1):
        name = chase.cell(r, 1).value
        if not name or str(name).startswith("◆"):
            continue
        hunt.cell(out_row, 1, name)
        hunt.cell(out_row, 2, chase.cell(r, 2).value)
        hunt.cell(out_row, 3, chase.cell(r, 3).value)
        hunt.cell(out_row, 4, chase.cell(r, 4).value)
        hunt.cell(out_row, 5, chase.cell(r, 5).value)
        hunt.cell(out_row, 6, chase.cell(r, 6).value)
        hunt.cell(out_row, 7, chase.cell(r, 7).value)
        hunt.cell(out_row, 8, chase.cell(r, 8).value)
        hunt.cell(out_row, 9, chase.cell(r, 14).value)  # verdict/notes
        out_row += 1
        added_chase += 1

    print(f"  ✓ Added {added_chase} chase cards to Hunt List sheet")

    dst.save(DEST)
    print(f"\n✓ Saved {DEST}")
    print(f"  Sheets now: {dst.sheetnames}")


if __name__ == "__main__":
    main()
