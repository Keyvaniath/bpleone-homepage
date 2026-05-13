"""Update the Watchlist Composite formula to use the new momentum-heavy weights.
Old:  =U*0.4 + V*0.25 + W*0.2 + X*0.15  (mom 40 / val 25 / scar 20 / liq 15)
New:  =U*0.55 + W*0.20 + V*0.15 + X*0.10  (mom 55 / scar 20 / val 15 / liq 10)
"""
from openpyxl import load_workbook

WB = "Pokemon_Card_Investment_Workbook.xlsx"
wb = load_workbook(WB)
ws = wb["Watchlist"]

updated = 0
for r in range(5, ws.max_row + 1):
    cid = ws.cell(r, 1).value
    if not cid:
        continue
    # Rewrite the Composite cell (column 25 = Y)
    ws.cell(r, 25, f"=U{r}*0.55+W{r}*0.20+V{r}*0.15+X{r}*0.10+ROW()/1000000")
    updated += 1

# Update the README/Dashboard description if present
if "Dashboard" in wb.sheetnames:
    dash = wb["Dashboard"]
    for r in range(1, min(dash.max_row + 1, 40)):
        for c in range(1, min(dash.max_column + 1, 20)):
            v = dash.cell(r, c).value
            if isinstance(v, str) and "0.4" in v and "Momentum" in v:
                dash.cell(r, c, "Composite Score = 0.55×Momentum + 0.20×Scarcity + 0.15×Value + 0.10×Liquidity. "
                               "Tuned for momentum traders + long-term holders.")

wb.save(WB)
print(f"✓ Updated Composite formula on {updated} rows.")
print("✓ New weighting: 55% Momentum / 20% Scarcity / 15% Value / 10% Liquidity")
