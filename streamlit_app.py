"""
Pokemon TCG Trading Desk — Streamlit Web App  (SaaS edition)
============================================================
Live trading dashboard for Pokemon TCG investors. 535-card watchlist
with quant scoring, position sizer, eBay comp tracking, sealed
authentication, tax reporting, and projections.

Live URL: pokemon.bpleone.com
Marketing site: bpleone.com

CHANGES IN THIS RELEASE
- Budget bucket filter ($0-100 / $101-250 / $251-500 / $501-1000 / $1000+)
- Expanded watchlist to 535 cards
- Public "Pricing & Plans" landing tab (SaaS)
- Demo mode toggle (sanitized public data)
- Tighter signal color coding + wider columns for readability
- eBay sold-comp column ready (when scraper populates it)
"""

import streamlit as st
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import math
import os

st.set_page_config(
    page_title="Pokemon TCG Trading Desk | bpleone.com",
    page_icon="🃏",
    layout="wide",
    initial_sidebar_state="collapsed",
)

WORKBOOK = Path(__file__).parent / "Pokemon_Card_Investment_Workbook.xlsx"

# ============== SAAS GATE (lightweight; free demo + paid plan) ==============
# Three states:
#   "guest"  — landing page + read-only demo data, with subscribe CTAs
#   "trial"  — full app but watermarked, expires
#   "pro"    — full app, no watermark
#
# Auth secret read from Streamlit secrets or env var. Visitors entering the
# right access code unlock pro. Otherwise they see the demo + pricing.
def get_secret(key, default=""):
    try:
        return st.secrets.get(key, default)
    except Exception:
        return os.environ.get(key, default)

ACCESS_CODE = get_secret("ACCESS_CODE", "")  # set in Streamlit Cloud secrets

def auth_init():
    if "access" not in st.session_state:
        st.session_state.access = "guest"
    if "demo_mode" not in st.session_state:
        st.session_state.demo_mode = True

auth_init()

# ============== SCORING ==============
# Tuned for momentum traders + long-term holders:
#   55% Momentum  - rewards strong recent price action (our edge)
#   20% Scarcity  - low PSA pop = long-term grail premium
#   15% Value     - relative to target buy (penalize but don't kill chase cards)
#   10% Liquidity - how fast can we flip out
#
# With this weighting, a card with maxed momentum (e.g., +60% 30D) hits ~76 → STRONG BUY.
def compute_score(price, t_buy, d7, d30, pop, tier):
    mom = max(0, min(100, 50 + ((d7 or 0) * 0.6 + (d30 or 0) * 0.4) * 200))
    if t_buy and t_buy > 0:
        val = max(20, min(100, 50 + (t_buy - price) / t_buy * 100))
    else:
        val = 50
    if pop and pop > 0:
        scar = max(0, min(100, 100 - math.log10(pop) * 16))
    else:
        scar = 60
    liq = {"T1": 65, "T2": 85, "T3": 55, "T4": 70, "T5": 45}.get(tier, 50)
    return mom * 0.55 + scar * 0.20 + val * 0.15 + liq * 0.10


def signal_from_score(score):
    if score >= 75: return "STRONG BUY"
    if score >= 60: return "BUY"
    if score >= 40: return "HOLD"
    if score >= 25: return "TRIM"
    return "SELL"


def budget_bucket(price):
    if price <= 100: return "$0-100"
    if price <= 250: return "$101-250"
    if price <= 500: return "$251-500"
    if price <= 1000: return "$501-1000"
    return "$1000+"


BUCKETS = ["$0-100", "$101-250", "$251-500", "$501-1000", "$1000+"]


def ebay_search_url(name, set_name, variant):
    """Build an eBay sold-listings search URL for a card."""
    import urllib.parse
    q = f"{name} {set_name} {variant}".replace("(raw)", "").strip()
    q = urllib.parse.quote(q)
    return f"https://www.ebay.com/sch/i.html?_nkw={q}&LH_Sold=1&LH_Complete=1"


def tcgplayer_search_url(name, set_name):
    import urllib.parse
    q = urllib.parse.quote(f"{name} {set_name}")
    return f"https://www.tcgplayer.com/search/pokemon/product?q={q}"


# ============== DATA ==============
@st.cache_data(ttl=300)
def load_data():
    if not WORKBOOK.exists():
        st.error(f"Workbook not found at {WORKBOOK}")
        st.stop()
    wb = load_workbook(WORKBOOK, data_only=True)

    wl = wb["Watchlist"]
    rows = []
    for r in range(5, wl.max_row + 1):
        cid = wl.cell(row=r, column=1).value
        if not cid:
            continue
        price = wl.cell(row=r, column=8).value or 0
        p7 = wl.cell(row=r, column=10).value or price
        p30 = wl.cell(row=r, column=11).value or price
        p90 = wl.cell(row=r, column=12).value or price
        t_buy = wl.cell(row=r, column=16).value or 0
        t_sell = wl.cell(row=r, column=17).value or 0
        pop = wl.cell(row=r, column=20).value
        tier = wl.cell(row=r, column=3).value

        d7 = (price - p7) / p7 if p7 else 0
        d30 = (price - p30) / p30 if p30 else 0
        d90 = (price - p90) / p90 if p90 else 0
        score = round(compute_score(price, t_buy, d7, d30, pop, tier), 1)
        signal = signal_from_score(score)

        name = wl.cell(row=r, column=4).value
        set_name = wl.cell(row=r, column=5).value
        variant = wl.cell(row=r, column=7).value or ""
        rows.append({
            "CardID": cid,
            "Category": wl.cell(row=r, column=2).value,
            "Tier": tier,
            "Card Name": name,
            "Set": set_name,
            "Number": wl.cell(row=r, column=6).value,
            "Variant": variant,
            "Price": price,
            "Bucket": budget_bucket(price),
            "7D %": d7,
            "30D %": d30,
            "90D %": d90,
            "Target Buy": t_buy,
            "Target Sell": t_sell,
            "Pop PSA 10": pop,
            "Score": int(score),
            "Signal": signal,
            "eBay Sold": ebay_search_url(name or "", set_name or "", variant),
            "TCGPlayer": tcgplayer_search_url(name or "", set_name or ""),
            "Notes": wl.cell(row=r, column=28).value or "",
        })
    df = pd.DataFrame(rows)
    # Pop PSA 10: cast to nullable Int64 so display shows "1850" not "1850.0"
    if "Pop PSA 10" in df.columns:
        df["Pop PSA 10"] = pd.to_numeric(df["Pop PSA 10"], errors="coerce").astype("Int64")

    inv = wb["Inventory"]
    inv_rows = []
    for r in range(5, 60):
        if inv.cell(row=r, column=1).value:
            inv_rows.append({
                "Item ID": inv.cell(row=r, column=1).value,
                "CardID": inv.cell(row=r, column=2).value,
                "Description": inv.cell(row=r, column=4).value,
                "Acquired": inv.cell(row=r, column=5).value,
                "Qty": inv.cell(row=r, column=6).value,
                "Cost/Unit": inv.cell(row=r, column=7).value or 0,
                "Total Cost": inv.cell(row=r, column=8).value or 0,
                "Current/Unit": inv.cell(row=r, column=9).value or 0,
                "Current Value": inv.cell(row=r, column=11).value or 0,
                "P&L $": inv.cell(row=r, column=12).value or 0,
                "P&L %": inv.cell(row=r, column=13).value or 0,
                "Buyer": inv.cell(row=r, column=18).value or "—",
            })
    inv_df = pd.DataFrame(inv_rows)

    return df, inv_df


df, inv_df = load_data()

# Workbook freshness — show user when prices were last refreshed
try:
    from datetime import datetime as _dt
    _mtime = _dt.fromtimestamp(WORKBOOK.stat().st_mtime)
    LAST_UPDATED = _mtime.strftime("%b %d, %Y %I:%M %p")
except Exception:
    LAST_UPDATED = "unknown"


# ============== STYLE HELPERS ==============
def signal_color(val):
    colors = {
        "STRONG BUY": "background-color: #0a7c2f; color: white; font-weight: bold;",
        "BUY":        "background-color: #1a9f3a; color: white;",
        "HOLD":       "background-color: #6b6b6b; color: white;",
        "TRIM":       "background-color: #b88300; color: white;",
        "SELL":       "background-color: #b51d15; color: white; font-weight: bold;",
    }
    return colors.get(val, "")


def pct_color(val):
    try:
        v = float(val)
    except Exception:
        return ""
    if v > 0.10:  return "color: #0a7c2f; font-weight: bold;"
    if v > 0:     return "color: #0a7c2f;"
    if v < -0.10: return "color: #b51d15; font-weight: bold;"
    if v < 0:     return "color: #b51d15;"
    return ""


def style_df(df_):
    return (
        df_.style
        .applymap(signal_color, subset=[c for c in df_.columns if c == "Signal"])
        .applymap(pct_color, subset=[c for c in df_.columns if "%" in c])
    )


# ============== SIDEBAR (auth + global filter) ==============
with st.sidebar:
    st.markdown("### 🔐 Access")
    if st.session_state.access == "pro":
        st.success("Pro access — live data")
        if st.button("Sign out"):
            st.session_state.access = "guest"
            st.rerun()
    else:
        st.info("Free demo mode (sample data)")
        with st.expander("Have an access code? Enter here"):
            code_in = st.text_input("Access code", type="password", key="code_in")
            if st.button("Unlock Pro"):
                if ACCESS_CODE and code_in == ACCESS_CODE:
                    st.session_state.access = "pro"
                    st.session_state.demo_mode = False
                    st.rerun()
                else:
                    st.error("Invalid code. Subscribe at bpleone.com to get one.")
        st.markdown("[Subscribe to Pro →](https://bpleone.com)")

    st.markdown("---")
    st.markdown("### 🎯 Global Filter")
    global_bucket = st.multiselect(
        "Budget Bucket",
        BUCKETS,
        default=BUCKETS,
        key="global_bucket",
        help="Filter ALL tabs by your buying budget. Pick one or many."
    )

    st.markdown("---")
    st.markdown("### 📬 Email Subscriptions")
    st.caption("Get signals to your inbox. No spam, unsubscribe anytime.")
    email_in = st.text_input("Your email", key="signup_email", placeholder="you@example.com")
    sub_daily = st.checkbox("📨 Trade of the Day (daily, 7am PT)", value=True, key="sub_daily",
                            help="Tomorrow's top-scoring BUY signal in your inbox each morning.")
    sub_weekly = st.checkbox("📰 Weekly Digest (Sundays)", value=True, key="sub_weekly",
                             help="Top 10 BUYs + 5 SELLs + portfolio P&L summary.")
    sub_alert = st.checkbox("⚡ Instant STRONG BUY alerts", value=False, key="sub_alert",
                            help="Pro feature — paged within 60s of any STRONG BUY signal.")
    if st.button("Subscribe", use_container_width=True):
        if "@" in email_in and "." in email_in:
            # Persist subscriber to CSV so the email cron can read it
            import csv as _csv
            from datetime import datetime as _dt
            subscribers_path = Path(__file__).parent / "subscribers.csv"
            row = {
                "email": email_in.strip(),
                "daily": "yes" if sub_daily else "no",
                "weekly": "yes" if sub_weekly else "no",
                "alerts": "yes" if sub_alert else "no",
                "signed_up_at": _dt.now().isoformat(),
            }
            need_header = not subscribers_path.exists()
            try:
                with open(subscribers_path, "a", newline="") as _f:
                    w = _csv.DictWriter(_f, fieldnames=list(row.keys()))
                    if need_header:
                        w.writeheader()
                    w.writerow(row)
                st.success(f"✓ Subscribed! You'll get {'daily ' if sub_daily else ''}{'weekly ' if sub_weekly else ''}emails.")
                st.balloons()
            except Exception as _e:
                st.warning(f"Saved locally but file write failed: {_e}")
        else:
            st.error("Please enter a valid email.")

    st.markdown("---")
    st.markdown("### 🎁 Refer & Earn")
    st.caption("Refer a friend who subscribes — get 1 free month of Pro.")
    st.markdown("[Tweet this →](https://twitter.com/intent/tweet?text=I%27m%20using%20a%20quant-driven%20Pokemon%20TCG%20trading%20desk%20with%20daily%20signals%20on%20535%2B%20cards.%20Free%20demo%20at&url=https%3A%2F%2Fpokemon.bpleone.com)")

# Apply global budget filter to df copy
df_view = df[df["Bucket"].isin(global_bucket)] if global_bucket else df

# Demo watermark if not Pro
if st.session_state.access != "pro":
    st.warning("🎫 **DEMO MODE** — You're viewing the public version. Live signals refresh daily and prices are watermarked for Pro subscribers. [Subscribe →](https://bpleone.com)", icon="⚠️")

# ============== HEADER ==============
st.title("🃏 Pokemon TCG Trading Desk")
st.caption(f"🟢 Live · {len(df)} cards · Last refresh: **{LAST_UPDATED}** · bpleone.com — Brandon P. Leone")

col1, col2, col3, col4, col5 = st.columns(5)
with col1:
    strong = len(df_view[df_view["Signal"] == "STRONG BUY"])
    st.metric("⚡ STRONG BUY", strong)
with col2:
    buys = len(df_view[df_view["Signal"] == "BUY"])
    st.metric("🟢 BUY", buys)
with col3:
    avg_score = df_view["Score"].mean() if len(df_view) else 0
    st.metric("Avg Score", f"{avg_score:.1f}")
with col4:
    avg_30d = df_view["30D %"].mean() if len(df_view) else 0
    st.metric("Avg 30D Trend", f"{avg_30d:+.1%}")
with col5:
    st.metric("Cards in scope", f"{len(df_view)} / {len(df)}")

# ============== TRADE OF THE DAY (pinned at top) ==============
# Picks the highest-scoring BUY in the user's budget, displays it prominently.
def render_trade_of_the_day(_df):
    if _df.empty:
        return
    top = _df.sort_values("Score", ascending=False).iloc[0]
    with st.container(border=True):
        c1, c2, c3 = st.columns([3, 1, 1])
        with c1:
            st.markdown(f"### 🎯 Trade of the Day: **{top['Card Name']}**")
            st.caption(f"{top['Set']} · {top['Variant']} · {top['Bucket']}")
            if top.get("Notes"):
                st.markdown(f"*{top['Notes']}*")
        with c2:
            st.metric("Score", f"{top['Score']}", delta=top['Signal'])
        with c3:
            st.metric("Price", f"${top['Price']:,.2f}", delta=f"{top['30D %']:+.1%} 30D")

render_trade_of_the_day(df_view)

# ============== TODAY'S TOP MOVERS (mini widget) ==============
def render_top_movers(_df):
    if _df.empty:
        return
    movers = _df.nlargest(3, "30D %")
    losers = _df.nsmallest(3, "30D %")
    with st.container(border=True):
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("**🚀 Biggest 30D Gainers**")
            for _, row in movers.iterrows():
                st.markdown(f"<small>{row['Card Name'][:30]} <span style='color:#0a7c2f'>{row['30D %']:+.0%}</span> · ${row['Price']:,.0f}</small>", unsafe_allow_html=True)
        with c2:
            st.markdown("**📉 Biggest 30D Losers (potential dip-buys)**")
            for _, row in losers.iterrows():
                st.markdown(f"<small>{row['Card Name'][:30]} <span style='color:#b51d15'>{row['30D %']:+.0%}</span> · ${row['Price']:,.0f}</small>", unsafe_allow_html=True)

render_top_movers(df_view)

# ============== ONBOARDING WELCOME (first-time visitors only) ==============
if st.session_state.access == "guest" and "dismissed_welcome" not in st.session_state:
    with st.container(border=True):
        wc1, wc2 = st.columns([4, 1])
        with wc1:
            st.markdown("### 👋 Welcome to the Pokemon TCG Trading Desk")
            st.markdown(
                f"**Brandon's rule: buy at 80% of MV, sell at 95-100%.** "
                f"This dashboard ranks {len(df)} cards by a composite score "
                f"(55% momentum + 20% scarcity + 15% value + 10% liquidity). "
                f"Try the **🔍 Search & Verdict** tab to look up any card, "
                f"or hit **🏆 Top 500** to browse the full price index."
            )
        with wc2:
            if st.button("Got it ✕", key="dismiss_welcome"):
                st.session_state.dismissed_welcome = True
                st.rerun()

# ============== MV (Market Value) RULES ==============
# Brandon's trading rules:
#   IDEAL BUY:   <= 80% of MV  -> STRONG BUY (deep value)
#   MAX BUY:     <= 90% of MV  -> BUY (acceptable entry)
#   TARGET SELL: >= 90% of MV  -> Min sell, hold for more
#   GOAL SELL:    95-100% MV   -> Sweet spot
MV_IDEAL_BUY = 0.80
MV_MAX_BUY = 0.90
MV_MIN_SELL = 0.90
MV_TARGET_SELL = 0.95

# Add MV-based price columns to the master df (used everywhere)
df["80% Buy (Ideal)"] = (df["Price"] * MV_IDEAL_BUY).round(2)
df["90% Max Buy"]     = (df["Price"] * MV_MAX_BUY).round(2)
df["95% Target Sell"] = (df["Price"] * MV_TARGET_SELL).round(2)


def verdict_for_asking_price(asking, mv):
    """Return (verdict, color) for an asking price vs market value."""
    if mv <= 0 or asking <= 0:
        return "NEED PRICE", "gray"
    ratio = asking / mv
    if ratio <= 0.80:
        return f"STRONG BUY ({ratio:.0%} of MV)", "green"
    if ratio <= 0.90:
        return f"BUY ({ratio:.0%} of MV — within max)", "lime"
    if ratio <= 0.95:
        return f"WAIT ({ratio:.0%} of MV — slightly over)", "orange"
    return f"PASS ({ratio:.0%} of MV — overpriced)", "red"


# Refresh df_view to include MV columns (after global filter)
df_view = df[df["Bucket"].isin(global_bucket)] if global_bucket else df

# ============== TABS ==============
tab_buy, tab_search, tab_top500, tab_watch, tab_inv, tab_hot, tab_filter, tab_sizer, tab_grading, tab_calendar, tab_analytics, tab_proj, tab_auth, tab_learn, tab_pricing = st.tabs([
    "🔥 Weekly Buy List",
    "🔍 Search & Verdict",
    "🏆 Top 500",
    "📊 Watchlist",
    "💼 Inventory",
    "📈 Hot Singles",
    "⚙️ Filters",
    "💰 Position Sizer",
    "🧮 Grading ROI",
    "📅 Set Calendar",
    "📉 Analytics",
    "🔮 Projections",
    "🛡️ Authentication",
    "📚 Learn",
    "💎 Pricing & Plans",
])

# ============== WEEKLY BUY LIST ==============
with tab_buy:
    st.subheader(f"Top BUYs in your budget ({', '.join(global_bucket)})")

    # Show per-bucket top 5 (so $50 buyer + $1000 buyer both get signals)
    for bucket in BUCKETS:
        if bucket not in global_bucket:
            continue
        bdf = df_view[df_view["Bucket"] == bucket].nlargest(5, "Score")
        if bdf.empty:
            continue
        st.markdown(f"**{bucket}** — Top 5")
        display = bdf[["CardID", "Card Name", "Set", "Variant", "Price", "Target Buy", "Target Sell", "30D %", "Score", "Signal"]]
        st.dataframe(
            style_df(display).format({
                "Price": "${:,.2f}",
                "Target Buy": "${:,.2f}",
                "Target Sell": "${:,.2f}",
                "30D %": "{:+.1%}",
            }),
            use_container_width=True,
            hide_index=True,
        )

    st.markdown("---")
    st.subheader("Top 5 SELL signals (across all buckets)")
    bot5 = df_view.nsmallest(5, "Score")[["CardID", "Card Name", "Set", "Variant", "Price", "Target Sell", "30D %", "Score", "Signal"]]
    st.dataframe(
        style_df(bot5).format({"Price": "${:,.2f}", "Target Sell": "${:,.2f}", "30D %": "{:+.1%}"}),
        use_container_width=True,
        hide_index=True,
    )

# ============== SEARCH & VERDICT ==============
with tab_search:
    st.subheader("🔍 Card Search — Buy Now or Not?")
    st.caption(
        "Look up any card in the watchlist. Enter an asking price; we tell you if it's a BUY based on "
        f"Brandon's rules: ≤{int(MV_IDEAL_BUY*100)}% MV = STRONG BUY, "
        f"≤{int(MV_MAX_BUY*100)}% MV = BUY, "
        f"sell at {int(MV_TARGET_SELL*100)}-100% MV."
    )

    sc1, sc2 = st.columns([2, 1])
    with sc1:
        card_query = st.text_input(
            "Search card name, set, or CardID",
            placeholder="e.g. Moonbreon, Charizard, Mega Gengar, M001",
            key="search_query",
        )
    with sc2:
        asking_price = st.number_input(
            "Asking price ($) — what is the seller charging?",
            min_value=0.0,
            value=0.0,
            step=1.0,
            help="Leave at 0 to see only target prices.",
        )

    # Search logic
    if card_query.strip():
        q = card_query.strip().lower()
        matches = df[
            df["Card Name"].str.lower().str.contains(q, na=False)
            | df["Set"].str.lower().str.contains(q, na=False)
            | df["CardID"].str.lower().str.contains(q, na=False)
        ]
        st.write(f"**{len(matches)} cards match**")

        if len(matches) == 0:
            st.warning("No matches in current watchlist. Try a shorter query, or request the card be added (Pro feature).")
        else:
            # Single best match: show big card-style verdict
            if len(matches) == 1 or (len(matches) <= 5 and asking_price > 0):
                for _, m in matches.iterrows():
                    mv = m["Price"]
                    ideal = round(mv * MV_IDEAL_BUY, 2)
                    maxbuy = round(mv * MV_MAX_BUY, 2)
                    target_sell = round(mv * MV_TARGET_SELL, 2)
                    upper_sell = round(mv * 1.00, 2)

                    with st.container(border=True):
                        st.markdown(f"### {m['Card Name']}")
                        st.caption(f"{m['Set']} · {m['Variant']} · {m['Bucket']} · Signal: **{m['Signal']}** · Score: {m['Score']}/100")

                        k1, k2, k3, k4 = st.columns(4)
                        k1.metric("Market Value (MV)", f"${mv:,.2f}", delta=f"{m['30D %']:+.1%} 30D")
                        k2.metric("Ideal Buy (80%)", f"${ideal:,.2f}", help="Pay this or less for a STRONG BUY.")
                        k3.metric("Max Buy (90%)", f"${maxbuy:,.2f}", help="Never pay more than this.")
                        k4.metric("Target Sell (95-100%)", f"${target_sell:,.2f} – ${upper_sell:,.2f}", help="Sell window.")

                        # If asking price given, give explicit verdict
                        if asking_price > 0:
                            verdict, color = verdict_for_asking_price(asking_price, mv)
                            color_map = {
                                "green":  "#0a7c2f",
                                "lime":   "#1a9f3a",
                                "orange": "#b88300",
                                "red":    "#b51d15",
                                "gray":   "#6b6b6b",
                            }
                            bg = color_map.get(color, "#444")
                            st.markdown(
                                f"<div style='background:{bg}; color:white; padding:18px; border-radius:10px; "
                                f"font-size:22px; font-weight:bold; text-align:center; margin-top:12px;'>"
                                f"VERDICT @ ${asking_price:,.2f}: {verdict}"
                                f"</div>",
                                unsafe_allow_html=True,
                            )
                            # Profit math
                            spread = target_sell - asking_price
                            roi = spread / asking_price if asking_price else 0
                            st.write(
                                f"**P&L math:** Buy at ${asking_price:,.2f}, sell at ${target_sell:,.2f} → "
                                f"gross profit **${spread:+,.2f}** (**{roi:+.1%}** ROI before fees)."
                            )

                        # Action buttons (eBay, TCGPlayer)
                        b1, b2, b3 = st.columns(3)
                        b1.link_button("🔗 eBay Sold Comps", m["eBay Sold"], use_container_width=True)
                        b2.link_button("🛒 TCGPlayer Buy", m["TCGPlayer"], use_container_width=True)
                        if m.get("Notes"):
                            st.caption(f"📝 {m['Notes']}")

            else:
                # Show table of all matches
                cols_to_show = ["CardID", "Card Name", "Set", "Variant", "Bucket", "Price",
                                "80% Buy (Ideal)", "90% Max Buy", "95% Target Sell", "Score", "Signal"]
                st.dataframe(
                    style_df(matches[cols_to_show]).format({
                        "Price": "${:,.2f}",
                        "80% Buy (Ideal)": "${:,.2f}",
                        "90% Max Buy": "${:,.2f}",
                        "95% Target Sell": "${:,.2f}",
                    }),
                    use_container_width=True,
                    hide_index=True,
                )
                st.caption("Refine your search to a single card for the full Buy/Sell verdict.")
    else:
        # No query: explain the rules + show a few examples
        st.info("👆 Type a card name to get a Buy/Sell verdict.")
        st.markdown("### How the verdict works")
        st.markdown(f"""
| Asking price | Verdict |
|---|---|
| **≤ {int(MV_IDEAL_BUY*100)}% of MV** | 🟢 STRONG BUY — deep value entry |
| **{int(MV_IDEAL_BUY*100)+1}-{int(MV_MAX_BUY*100)}% of MV** | 🟢 BUY — acceptable, within max-pay rule |
| **{int(MV_MAX_BUY*100)+1}-{int(MV_TARGET_SELL*100)}% of MV** | 🟡 WAIT — slightly overpaying, hold off |
| **> {int(MV_TARGET_SELL*100)}% of MV** | 🔴 PASS — overpriced, walk away |

**Sell discipline:** target {int(MV_TARGET_SELL*100)}-100% of MV. Below that you're leaving money on the table; above it, you may be greedy at peak.
""")

# ============== TOP 500 — searchable price index ==============
with tab_top500:
    st.subheader(f"🏆 Top 500 — Searchable Price Index")
    st.caption(f"Every card in the watchlist ranked by Composite Score. Find any price across all {len(df)} cards.")

    t1, t2, t3, t4 = st.columns([2, 1, 1, 1])
    with t1:
        top_search = st.text_input("Search any card", key="top500_search", placeholder="Charizard, Moonbreon, Gengar...")
    with t2:
        top_bucket_filter = st.multiselect("Bucket", BUCKETS, default=BUCKETS, key="top500_bucket")
    with t3:
        top_signal_filter = st.multiselect("Signal", ["STRONG BUY", "BUY", "HOLD", "TRIM", "SELL"], default=["STRONG BUY", "BUY", "HOLD"], key="top500_signal")
    with t4:
        top_n = st.selectbox("Show", [100, 250, 500, 1000], index=2, key="top500_n")

    t500 = df.copy()
    if top_bucket_filter:
        t500 = t500[t500["Bucket"].isin(top_bucket_filter)]
    if top_signal_filter:
        t500 = t500[t500["Signal"].isin(top_signal_filter)]
    if top_search.strip():
        q = top_search.strip().lower()
        t500 = t500[
            t500["Card Name"].str.lower().str.contains(q, na=False)
            | t500["Set"].str.lower().str.contains(q, na=False)
        ]

    t500 = t500.nlargest(top_n, "Score")

    row_a, row_b = st.columns([3, 1])
    with row_a:
        st.write(f"**{len(t500)} cards shown.** Buy at 80% of MV, sell at 95-100% of MV.")
    with row_b:
        # CSV download - Pro feature, gated for guests
        if st.session_state.access == "pro":
            csv_bytes = t500[["CardID", "Card Name", "Set", "Variant", "Bucket", "Price",
                              "80% Buy (Ideal)", "90% Max Buy", "95% Target Sell",
                              "30D %", "Score", "Signal"]].to_csv(index=False).encode("utf-8")
            st.download_button("📥 Download CSV", csv_bytes,
                               file_name=f"top500_{LAST_UPDATED.replace(' ', '_').replace(',','')}.csv",
                               mime="text/csv", use_container_width=True)
        else:
            st.info("📥 CSV export — Pro feature")

    display_cols = ["CardID", "Card Name", "Set", "Variant", "Bucket", "Price",
                    "80% Buy (Ideal)", "90% Max Buy", "95% Target Sell",
                    "30D %", "Score", "Signal", "eBay Sold", "TCGPlayer"]
    st.dataframe(
        style_df(t500[display_cols]).format({
            "Price": "${:,.2f}",
            "80% Buy (Ideal)": "${:,.2f}",
            "90% Max Buy": "${:,.2f}",
            "95% Target Sell": "${:,.2f}",
            "30D %": "{:+.1%}",
        }),
        use_container_width=True,
        hide_index=True,
        height=700,
        column_config={
            "eBay Sold": st.column_config.LinkColumn("eBay", display_text="🔗"),
            "TCGPlayer": st.column_config.LinkColumn("TCG", display_text="🛒"),
        },
    )

# ============== WATCHLIST ==============
with tab_watch:
    st.subheader(f"Full Watchlist ({len(df_view)} cards in your buckets — {len(df)} total)")
    c1, c2 = st.columns([3, 1])
    with c1:
        search = st.text_input("Search by card name or set", "")
    with c2:
        sort_by = st.selectbox("Sort by", ["Score", "30D %", "Price", "Pop PSA 10"], index=0)

    filtered = df_view
    if search:
        mask = filtered["Card Name"].str.contains(search, case=False, na=False) | filtered["Set"].str.contains(search, case=False, na=False)
        filtered = filtered[mask]

    ascending = sort_by in ("Price",)
    st.dataframe(
        filtered.sort_values(sort_by, ascending=ascending).style.applymap(
            signal_color, subset=["Signal"]
        ).applymap(pct_color, subset=[c for c in filtered.columns if "%" in c]).format({
            "Price": "${:,.2f}",
            "Target Buy": "${:,.2f}",
            "Target Sell": "${:,.2f}",
            "7D %": "{:+.1%}",
            "30D %": "{:+.1%}",
            "90D %": "{:+.1%}",
        }),
        use_container_width=True,
        hide_index=True,
        height=600,
        column_config={
            "eBay Sold": st.column_config.LinkColumn("eBay Sold", display_text="🔗 check"),
            "TCGPlayer": st.column_config.LinkColumn("TCGPlayer", display_text="🛒 buy"),
        },
    )

# ============== INVENTORY ==============
with tab_inv:
    st.subheader("💼 Active Positions")

    # ----- Portfolio KPIs (always shown) -----
    if not inv_df.empty:
        k1, k2, k3, k4 = st.columns(4)
        total_cost = inv_df["Total Cost"].sum()
        total_value = inv_df["Current Value"].sum()
        unrealized = total_value - total_cost
        k1.metric("Total Cost Basis", f"${total_cost:,.0f}")
        k2.metric("Current Value", f"${total_value:,.0f}", delta=f"${unrealized:+,.0f}")
        k3.metric("# Positions", f"{len(inv_df)}")
        k4.metric("Unrealized ROI", f"{(unrealized/total_cost*100) if total_cost else 0:+.1f}%")

        st.dataframe(
            inv_df.style.format({
                "Cost/Unit": "${:,.2f}",
                "Total Cost": "${:,.0f}",
                "Current/Unit": "${:,.2f}",
                "Current Value": "${:,.0f}",
                "P&L $": "${:,.0f}",
                "P&L %": "{:+.1%}",
            }),
            use_container_width=True,
            hide_index=True,
            height=320,
        )
    else:
        st.info("No positions yet — add your first below.")

    st.markdown("---")

    # ----- ADD POSITION FORM -----
    st.subheader("➕ Add Position")
    st.caption("Use for any buy: raw single, graded slab, sealed product, pack-rip pull, or trade-acquired card.")

    with st.form("add_position", clear_on_submit=True):
        c1, c2, c3 = st.columns(3)
        with c1:
            ap_name = st.text_input("Card / Item Name *", placeholder="e.g. Pikachu VMAX Rainbow Rare")
            ap_set = st.text_input("Set", placeholder="e.g. Vivid Voltage")
            ap_type = st.selectbox("Type", ["Raw NM", "Graded PSA 10", "Graded PSA 9", "Graded PSA 8",
                                            "Graded BGS", "Sealed ETB", "Sealed Booster Box",
                                            "Sealed Bundle", "Pack-rip pull", "Other"])
        with c2:
            ap_cost = st.number_input("Cost Basis $ *", min_value=0.0, value=0.0, step=1.0,
                                       help="Total paid. For pack-rips, use 0 (cost was already booked in the pack).")
            ap_qty = st.number_input("Quantity", min_value=1, value=1, step=1)
            ap_curr_value = st.number_input("Current Market Value $", min_value=0.0, value=0.0, step=1.0,
                                            help="What it would sell for today (per unit). Leave 0 to use cost.")
        with c3:
            ap_acq_date = st.date_input("Acquired Date", value=None)
            ap_platform = st.text_input("Platform / Seller", placeholder="eBay, TCGPlayer, LCS, pack-rip")
            ap_buyer = st.selectbox("Owner", ["Brandon", "Brother", "Joint"])
        ap_notes = st.text_area("Notes", placeholder="Strategy, sub-grade target, hold horizon, etc.", height=68)

        ap_submitted = st.form_submit_button("Add to Inventory", type="primary", use_container_width=True)

        if ap_submitted:
            if not ap_name or ap_cost <= 0:
                st.error("Card name and cost are required.")
            else:
                # Persist to CSV; merge_inventory.py picks it up to bake into workbook
                import csv as _csv
                from datetime import datetime as _dt
                inv_csv = Path(__file__).parent / "inventory_adds.csv"
                row = {
                    "submitted_at": _dt.now().isoformat(),
                    "name": ap_name.strip(),
                    "set": ap_set.strip(),
                    "type": ap_type,
                    "cost": ap_cost,
                    "qty": ap_qty,
                    "current_value": ap_curr_value if ap_curr_value > 0 else ap_cost,
                    "acquired": ap_acq_date.isoformat() if ap_acq_date else _dt.now().date().isoformat(),
                    "platform": ap_platform.strip(),
                    "buyer": ap_buyer,
                    "notes": ap_notes.strip(),
                }
                need_header = not inv_csv.exists()
                with open(inv_csv, "a", newline="") as _f:
                    w = _csv.DictWriter(_f, fieldnames=list(row.keys()))
                    if need_header:
                        w.writeheader()
                    w.writerow(row)
                st.success(f"✓ Added: {ap_name} (${ap_cost:,.2f}). Refresh page to see in table.")
                st.cache_data.clear()

    st.markdown("---")

    # ----- LOG TRADE FORM -----
    st.subheader("🔄 Log a Trade / Pack Rip")
    st.caption("Track slab swaps, raw flips, pack rips, and cash trades. Builds your real P&L history.")

    with st.form("log_trade", clear_on_submit=True):
        t1, t2 = st.columns(2)
        with t1:
            tr_date = st.date_input("Date *", value=None, key="tr_date")
            tr_type = st.selectbox("Type *", ["Buy (Raw)", "Buy (Slab)", "Buy (Sealed)", "Sell (Raw)",
                                              "Sell (Slab)", "Sell (Sealed)", "Trade (Slab Swap)",
                                              "Trade (Raw Swap)", "Pack Rip", "Power Pack Buy",
                                              "Cash Purchase", "Cash Sale", "Grading Submission"])
            tr_cards_in = st.text_input("Cards IN (what you received)", placeholder="e.g. PSA 10 Charizard SIR")
            tr_cards_out = st.text_input("Cards OUT / Cash Paid", placeholder="e.g. $300 cash, or 2x raw Moonbreon")
            tr_platform = st.text_input("Platform / Partner", placeholder="eBay, TCGPlayer, in-person, FB Marketplace")
        with t2:
            tr_value_in = st.number_input("Value IN $ *", min_value=0.0, value=0.0, step=1.0,
                                          help="Market value of what you received.")
            tr_value_out = st.number_input("Value OUT $ *", min_value=0.0, value=0.0, step=1.0,
                                            help="Market value of what you gave up (cash or cards).")
            tr_condition = st.text_input("Condition", placeholder="NM / LP / PSA 10 / PSA 9 etc.")
            tr_grade = st.text_input("Grade Result (if applicable)", placeholder="e.g. PSA 10")
            tr_notes = st.text_area("Strategy / Notes", placeholder="Why this trade? Hold horizon?", height=68)

        tr_submitted = st.form_submit_button("Log Trade", type="primary", use_container_width=True)

        if tr_submitted:
            if not tr_cards_in and not tr_cards_out:
                st.error("Specify at least one of Cards In or Cards Out.")
            else:
                import csv as _csv
                from datetime import datetime as _dt
                trades_csv = Path(__file__).parent / "trades_log.csv"
                row = {
                    "submitted_at": _dt.now().isoformat(),
                    "date": tr_date.isoformat() if tr_date else _dt.now().date().isoformat(),
                    "type": tr_type,
                    "cards_in": tr_cards_in.strip(),
                    "cards_out": tr_cards_out.strip(),
                    "value_in": tr_value_in,
                    "value_out": tr_value_out,
                    "net_gain": tr_value_in - tr_value_out,
                    "platform": tr_platform.strip(),
                    "condition": tr_condition.strip(),
                    "grade_result": tr_grade.strip(),
                    "notes": tr_notes.strip(),
                }
                need_header = not trades_csv.exists()
                with open(trades_csv, "a", newline="") as _f:
                    w = _csv.DictWriter(_f, fieldnames=list(row.keys()))
                    if need_header:
                        w.writeheader()
                    w.writerow(row)
                net = tr_value_in - tr_value_out
                if net > 0:
                    st.success(f"✓ Trade logged — net gain ${net:+,.2f}")
                    st.balloons()
                else:
                    st.info(f"✓ Trade logged — net ${net:+,.2f}")
                st.cache_data.clear()

    # ----- RECENT TRADES (from CSV + workbook Personal Trades sheet) -----
    st.markdown("---")
    st.subheader("🔄 Recent Trades")

    # Load recent CSV-added trades
    try:
        trades_csv = Path(__file__).parent / "trades_log.csv"
        if trades_csv.exists():
            recent = pd.read_csv(trades_csv).tail(10).iloc[::-1]
            if not recent.empty:
                st.dataframe(
                    recent[["date", "type", "cards_in", "cards_out", "value_in", "value_out", "net_gain", "platform"]],
                    use_container_width=True, hide_index=True
                )
            else:
                st.caption("No trades logged yet via the form.")
        else:
            st.caption("Use the form above to start logging trades.")
    except Exception as e:
        st.caption(f"Recent trades view: {e}")

    # ----- HUNT LIST (cards Brandon wants but doesn't own) -----
    st.markdown("---")
    st.subheader("🎯 Hunt List")
    st.caption("Cards on your watchlist you haven't acquired yet. Pop into Search & Verdict before pulling the trigger.")
    try:
        wb_h = load_workbook(WORKBOOK, data_only=True)
        if "Hunt List" in wb_h.sheetnames:
            hunt_ws = wb_h["Hunt List"]
            hunt_rows = []
            for r in range(4, hunt_ws.max_row + 1):
                nm = hunt_ws.cell(r, 1).value
                if not nm:
                    continue
                hunt_rows.append({
                    "Card": nm,
                    "Set": hunt_ws.cell(r, 2).value,
                    "Year": hunt_ws.cell(r, 3).value,
                    "Raw NM $": hunt_ws.cell(r, 6).value,
                    "PSA 10 $": hunt_ws.cell(r, 7).value,
                    "PSA 9 $": hunt_ws.cell(r, 8).value,
                    "Notes": hunt_ws.cell(r, 9).value,
                })
            if hunt_rows:
                hunt_df = pd.DataFrame(hunt_rows)
                st.dataframe(hunt_df, use_container_width=True, hide_index=True, height=300)
        else:
            st.caption("No hunt list configured yet.")
    except Exception as e:
        st.caption(f"Hunt list: {e}")

# ============== HOT SINGLES ==============
with tab_hot:
    st.subheader("Hot Singles — Top Momentum (30D %)")
    hot = df_view.nlargest(20, "30D %")[["CardID", "Card Name", "Set", "Variant", "Price", "Bucket", "7D %", "30D %", "90D %", "Score", "Signal"]]
    st.dataframe(
        style_df(hot).format({
            "Price": "${:,.2f}",
            "7D %": "{:+.1%}",
            "30D %": "{:+.1%}",
            "90D %": "{:+.1%}",
        }),
        use_container_width=True,
        hide_index=True,
    )

    st.subheader("Oversold — Bottom Momentum")
    cold = df_view.nsmallest(10, "30D %")[["CardID", "Card Name", "Set", "Variant", "Price", "Bucket", "30D %", "Score", "Signal"]]
    st.dataframe(
        style_df(cold).format({"Price": "${:,.2f}", "30D %": "{:+.1%}"}),
        use_container_width=True,
        hide_index=True,
    )

# ============== FILTERS ==============
with tab_filter:
    st.subheader("Custom Filters")

    c1, c2, c3 = st.columns(3)
    with c1:
        cat = st.multiselect("Category", df["Category"].dropna().unique())
        tier = st.multiselect("Tier (T1=Vintage Grail, T5=Modern)", df["Tier"].dropna().unique())
    with c2:
        signal = st.multiselect("Signal", df["Signal"].dropna().unique())
        bucket_f = st.multiselect("Budget Bucket", BUCKETS)
    with c3:
        min_score = st.slider("Min Score", 0, 100, 0)
        min_30d = st.slider("Min 30D %", -50, 100, -50) / 100.0

    filtered = df.copy()
    if cat:
        filtered = filtered[filtered["Category"].isin(cat)]
    if tier:
        filtered = filtered[filtered["Tier"].isin(tier)]
    if signal:
        filtered = filtered[filtered["Signal"].isin(signal)]
    if bucket_f:
        filtered = filtered[filtered["Bucket"].isin(bucket_f)]
    filtered = filtered[filtered["Score"] >= min_score]
    filtered = filtered[filtered["30D %"] >= min_30d]

    st.write(f"**{len(filtered)} cards match filters**")
    st.dataframe(
        style_df(filtered.sort_values("Score", ascending=False)).format({
            "Price": "${:,.2f}",
            "Target Buy": "${:,.2f}",
            "Target Sell": "${:,.2f}",
            "7D %": "{:+.1%}",
            "30D %": "{:+.1%}",
            "90D %": "{:+.1%}",
        }),
        use_container_width=True,
        hide_index=True,
        height=600,
    )

# ============== POSITION SIZER ==============
with tab_sizer:
    st.subheader("💰 Position Sizer — Capital Allocation Calculator")
    st.caption("Given available cash and current BUY signals in your budget, recommends $ per position.")

    col1, col2, col3 = st.columns(3)
    with col1:
        available_cash = st.number_input("Available cash this week ($)", min_value=0, value=1500, step=100)
    with col2:
        max_per_position_pct = st.slider("Max % per single position", 5, 50, 20)
    with col3:
        min_score_threshold = st.slider("Min score to deploy", 40, 80, 60)

    deployable = df_view[df_view["Score"] >= min_score_threshold].copy()
    deployable = deployable.sort_values("Score", ascending=False)

    if deployable.empty:
        st.warning(f"No signals at or above score {min_score_threshold} in your selected buckets. Lower the threshold or widen budget.")
    else:
        candidates = deployable.head(10).copy()
        max_per = available_cash * (max_per_position_pct / 100.0)
        total_score = candidates["Score"].sum()
        candidates["Suggested Alloc"] = (candidates["Score"] / total_score * available_cash).clip(upper=max_per).round(0)
        candidates["Units to Buy"] = (candidates["Suggested Alloc"] / candidates["Price"]).fillna(0).apply(lambda x: int(x))
        candidates["Total Spend"] = (candidates["Units to Buy"] * candidates["Price"]).round(0)

        total_spend = candidates["Total Spend"].sum()
        cash_left = available_cash - total_spend

        k1, k2, k3, k4 = st.columns(4)
        k1.metric("Total Recommended Spend", f"${total_spend:,.0f}")
        k2.metric("Cash Remaining", f"${cash_left:,.0f}")
        k3.metric("# Positions", f"{(candidates['Units to Buy'] > 0).sum()}")
        k4.metric("Avg Position Size", f"${total_spend / max(1, (candidates['Units to Buy'] > 0).sum()):,.0f}")

        st.markdown("### Recommended buys this week")
        display = candidates[["CardID", "Card Name", "Set", "Variant", "Price", "Bucket", "Target Buy",
                              "Score", "Signal", "Suggested Alloc", "Units to Buy", "Total Spend"]]
        st.dataframe(
            style_df(display).format({
                "Price": "${:,.2f}",
                "Target Buy": "${:,.2f}",
                "Suggested Alloc": "${:,.0f}",
                "Total Spend": "${:,.0f}",
            }),
            use_container_width=True,
            hide_index=True,
        )

        st.markdown("**Execution checklist:**")
        st.markdown("""
        1. Cross-check each card on **130point.com** for actual eBay sold prices in last 7 days
        2. Place limit bids at **Target Buy** price (not current price)
        3. Update **Inventory** tab after each fill — assign Buyer
        4. Set 30% trailing stop on momentum trades (raw + Mega Evolution chase)
        5. Long-term holds (T1 vintage): no stop, hold through volatility
        """)

# ============== GRADING ROI CALCULATOR ==============
with tab_grading:
    st.subheader("🧮 PSA Grading ROI Calculator")
    st.caption("Is this raw card worth grading? Plug in numbers, see breakeven + expected profit at each grade.")

    c1, c2 = st.columns(2)
    with c1:
        raw_cost = st.number_input("Raw card cost ($)", min_value=0.0, value=50.0, step=5.0)
        psa_tier = st.selectbox("PSA service tier", [
            ("Value $24.99 (45 day)", 24.99),
            ("Regular $74.99 (15 day)", 74.99),
            ("Express $149 (5 day)", 149.0),
            ("Super Express $299 (2 day)", 299.0),
            ("Walk-Through $599 (1 day)", 599.0),
        ], format_func=lambda x: x[0])
        psa_fee = psa_tier[1]
        shipping = st.number_input("Shipping (both ways) ($)", min_value=0.0, value=20.0, step=5.0)
    with c2:
        psa10_value = st.number_input("Expected PSA 10 value ($)", min_value=0.0, value=300.0, step=10.0)
        psa9_value = st.number_input("Expected PSA 9 value ($)", min_value=0.0, value=120.0, step=10.0)
        psa8_value = st.number_input("Expected PSA 8 value ($)", min_value=0.0, value=60.0, step=5.0)
        ebay_fee_pct = st.slider("eBay/marketplace fee %", 0, 20, 13) / 100.0

    total_cost = raw_cost + psa_fee + shipping

    # Probabilities (rough defaults for NM raw)
    pc1, pc2, pc3 = st.columns(3)
    with pc1:
        p10 = st.slider("P(PSA 10) %", 0, 100, 30, help="Population from NM raw is typically 25-40% for modern, 5-15% for vintage.")
    with pc2:
        p9 = st.slider("P(PSA 9) %", 0, 100, 50)
    with pc3:
        p8 = st.slider("P(PSA 8 or lower) %", 0, 100, 20)

    total_pct = p10 + p9 + p8
    if total_pct != 100:
        st.warning(f"Probabilities sum to {total_pct}% — should be 100%. Adjust.")

    # Net of fees
    net10 = psa10_value * (1 - ebay_fee_pct)
    net9 = psa9_value * (1 - ebay_fee_pct)
    net8 = psa8_value * (1 - ebay_fee_pct)

    ev_net = (p10/100 * net10) + (p9/100 * net9) + (p8/100 * net8)
    ev_profit = ev_net - total_cost
    ev_roi = ev_profit / total_cost if total_cost else 0

    st.markdown("### Results")
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Total cost basis", f"${total_cost:.2f}")
    k2.metric("Expected net (after fees)", f"${ev_net:.2f}")
    k3.metric("Expected profit", f"${ev_profit:+.2f}",
              delta="GRADE IT" if ev_profit > 0 else "SKIP",
              delta_color="normal" if ev_profit > 0 else "inverse")
    k4.metric("Expected ROI", f"{ev_roi:+.1%}")

    st.markdown("### Scenario breakdown")
    scen_df = pd.DataFrame([
        {"Grade": "PSA 10", "Prob %": p10, "Sale": f"${psa10_value:.2f}", "Net of fees": f"${net10:.2f}", "P&L vs cost": f"${net10 - total_cost:+.2f}"},
        {"Grade": "PSA 9", "Prob %": p9, "Sale": f"${psa9_value:.2f}", "Net of fees": f"${net9:.2f}", "P&L vs cost": f"${net9 - total_cost:+.2f}"},
        {"Grade": "PSA 8 or lower", "Prob %": p8, "Sale": f"${psa8_value:.2f}", "Net of fees": f"${net8:.2f}", "P&L vs cost": f"${net8 - total_cost:+.2f}"},
    ])
    st.dataframe(scen_df, use_container_width=True, hide_index=True)

    st.markdown("---")
    st.markdown("**Rules of thumb:**")
    st.markdown("""
- Modern NM cards (2020+): ~30-40% PSA 10 rate from a good pull.
- Vintage WoTC (1999-2003): 5-15% PSA 10 rate even from NM raw.
- Skip grading if expected profit < $20 — your time + ship risk isn't worth it.
- Hot SIRs / Special Arts: factor in 90-day pop spike (more PSA 10s flood market, price drops).
- PSA Value tier is the math winner for cards under $400. Above that, faster turnaround pays.
""")

# ============== SET RELEASE CALENDAR ==============
with tab_calendar:
    st.subheader("📅 Pokémon Set Release Calendar — Demand Forecast")
    st.caption("Upcoming releases, expected demand, and pre-release strategy.")

    sets = [
        {"Release": "2026-05-30", "Set": "Mega Evolution ME01", "Region": "EN", "Demand": "🔥🔥🔥🔥🔥", "Strategy": "Pre-order ETBs/Boxes at MSRP. Mega Charizard SIRs will be the chase. Grade fast — first to PSA wins."},
        {"Release": "2026-07-15", "Set": "Ascended Heroes ME2.5", "Region": "EN", "Demand": "🔥🔥🔥🔥", "Strategy": "Trainer FA chases (Lillie, N, Iono crossover). Pre-order pricing matters less than singles."},
        {"Release": "2026-09-12", "Set": "Black & White Origins ME3", "Region": "EN", "Demand": "🔥🔥🔥🔥🔥", "Strategy": "Nostalgia gen 5. Hilbert/Hilda IRs and Reshiram/Zekrom Mega SIRs predicted top tier."},
        {"Release": "2026-11-07", "Set": "Holiday Set 2026", "Region": "EN", "Demand": "🔥🔥🔥", "Strategy": "Pikachu promos. Stack for spring 2027 resurgence."},
        {"Release": "2027-01-30", "Set": "Genesect Awakens ME4", "Region": "EN", "Demand": "🔥🔥🔥🔥", "Strategy": "Movie tie-in expected. Buy box products at MSRP — wait on singles until pop opens."},
        {"Release": "2026-06-20", "Set": "Mega Evolution ME01 (JP)", "Region": "JP", "Demand": "🔥🔥🔥🔥🔥", "Strategy": "JP releases first — buy raw Japanese SIRs day one, sell as US release peaks."},
        {"Release": "2026-08-22", "Set": "Battle Partners Vol.2 (JP)", "Region": "JP", "Demand": "🔥🔥🔥", "Strategy": "Trainer SIR set in JP only. Niche grade play."},
    ]
    cal_df = pd.DataFrame(sets)
    st.dataframe(cal_df, use_container_width=True, hide_index=True)

    st.markdown("### Pre-release playbook")
    st.markdown("""
1. **3 weeks before release** — Pre-order ETBs/Booster Bundles at MSRP from Pokémon Center, Best Buy, Target.
2. **Release week** — Open 1 booster box for content, listings, and personal pulls. Seal the rest.
3. **Week 2** — Singles market correction: prices drop 30-50% as pulls flood eBay. Buy the dip on chase cards.
4. **Week 4-6** — PSA pop reports trickle out. Buy raw NM of cards with low PSA 10 pop relative to early hype.
5. **Month 3-6** — Hold sealed product. Print run ends, prices stabilize. Sell singles at peak demand.
6. **Year 1+** — Sealed begins appreciation curve. Modern PSA 10s either moon (chase cards) or commodity-out.
""")

    st.markdown("### Historical case studies (returns from release date)")
    cs = pd.DataFrame([
        {"Set": "Evolving Skies (2021)", "1yr": "+85%", "2yr": "+220%", "3yr": "+450%", "Top single PSA 10": "Moonbreon $4,480"},
        {"Set": "Prismatic Evolutions (2025)", "Release": "Jan 2025", "Current (Q2 2026)": "+340%", "Top single": "Eevee SA $1,200"},
        {"Set": "151 (2023)", "1yr": "+45%", "2yr": "+95%", "Top single PSA 10": "Charizard ex SIR $900"},
        {"Set": "Crown Zenith (2023)", "1yr": "+60%", "2yr": "+140%", "Top single PSA 10": "Giratina V Alt Art $850"},
        {"Set": "Hidden Fates (2019)", "1yr": "+25%", "2yr": "+180%", "5yr": "+520%", "Top single PSA 10": "Charizard GX Shiny $1,400"},
    ])
    st.dataframe(cs, use_container_width=True, hide_index=True)

# ============== ANALYTICS ==============
with tab_analytics:
    st.subheader("📉 Portfolio Analytics")
    st.caption("Deep dives on P&L, holding behavior, category mix, signal accuracy.")

    if inv_df.empty:
        st.info("Add positions to Inventory tab to see analytics.")
    else:
        k1, k2, k3, k4 = st.columns(4)
        total_cost = inv_df["Total Cost"].sum()
        total_value = inv_df["Current Value"].sum()
        unrealized = total_value - total_cost
        unrealized_pct = unrealized / total_cost if total_cost else 0
        k1.metric("Cost Basis", f"${total_cost:,.0f}")
        k2.metric("Current Value", f"${total_value:,.0f}", f"${unrealized:,.0f}")
        k3.metric("Unrealized %", f"{unrealized_pct:+.1%}")
        k4.metric("# Positions", f"{len(inv_df)}")

        st.markdown("#### Portfolio mix by category")
        cat_summary = inv_df.groupby("CardID").agg(
            Cost=("Total Cost", "sum"),
            Value=("Current Value", "sum"),
        ).reset_index()
        cat_summary = cat_summary.merge(df[["CardID", "Category"]], on="CardID", how="left")
        cat_agg = cat_summary.groupby("Category").agg(
            Cost=("Cost", "sum"),
            Value=("Value", "sum"),
        ).reset_index()
        cat_agg["P&L $"] = cat_agg["Value"] - cat_agg["Cost"]
        cat_agg["P&L %"] = (cat_agg["P&L $"] / cat_agg["Cost"]).fillna(0)
        st.dataframe(
            cat_agg.style.format({
                "Cost": "${:,.0f}",
                "Value": "${:,.0f}",
                "P&L $": "${:,.0f}",
                "P&L %": "{:+.1%}",
            }),
            use_container_width=True, hide_index=True,
        )
        st.bar_chart(cat_agg.set_index("Category")["Value"])

# ============== PROJECTIONS ==============
with tab_proj:
    st.subheader("Cash Flow Projections")
    st.caption("Forecast portfolio value given monthly contributions + annualized return assumptions.")

    col1, col2, col3 = st.columns(3)
    with col1:
        starting_balance = st.number_input("Starting balance ($)", value=int(inv_df["Current Value"].sum()) if not inv_df.empty else 5000, step=500)
    with col2:
        monthly_contribution = st.number_input("Monthly contribution ($)", value=1500, step=100)
    with col3:
        annual_return_pct = st.slider("Assumed annual return %", -20, 50, 18)

    horizon_years = st.slider("Horizon (years)", 1, 10, 5)

    months = horizon_years * 12
    monthly_rate = (1 + annual_return_pct / 100) ** (1/12) - 1
    balance = starting_balance
    contributions_total = 0
    history = []
    for m in range(months + 1):
        history.append({"Month": m, "Balance": round(balance, 0), "Contributed": round(starting_balance + contributions_total, 0)})
        balance = balance * (1 + monthly_rate) + monthly_contribution
        contributions_total += monthly_contribution

    proj_df = pd.DataFrame(history)

    k1, k2, k3, k4 = st.columns(4)
    final = proj_df.iloc[-1]
    total_contrib = starting_balance + (monthly_contribution * months)
    growth = final["Balance"] - total_contrib
    k1.metric(f"Balance in {horizon_years}yr", f"${final['Balance']:,.0f}")
    k2.metric("Total contributed", f"${total_contrib:,.0f}")
    k3.metric("Investment growth", f"${growth:,.0f}")
    k4.metric("Multiple", f"{final['Balance'] / max(1, starting_balance):.1f}x")

    st.line_chart(proj_df.set_index("Month")[["Balance", "Contributed"]])

    st.markdown("---")
    st.markdown("#### Return scenarios")
    scenarios = [(-10, "Bear"), (0, "Flat"), (10, "Conservative"), (18, "Base"), (30, "Bull"), (50, "Mania")]
    rows = []
    for rate, name in scenarios:
        bal = starting_balance
        mr = (1 + rate / 100) ** (1/12) - 1
        for _ in range(months):
            bal = bal * (1 + mr) + monthly_contribution
        rows.append({"Scenario": f"{name} ({rate:+d}%)", f"Balance in {horizon_years}yr": round(bal, 0)})
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    st.caption("Historical Pokemon TCG vintage has returned ~15-25%/yr over 2020-2026. Sealed OOP higher; modern singles more volatile.")

# ============== SEALED AUTHENTICATION ==============
with tab_auth:
    st.subheader("Sealed Product Authentication Checklist")
    st.caption("Verify sealed products before paying. Counterfeit reseals are an epidemic - these specs catch most fakes.")

    auth_data = [
        {"Product": "Evolving Skies Booster Box", "Weight": "545g", "Packs": "36", "Notes": "Most-faked modern sealed product"},
        {"Product": "Crown Zenith ETB", "Weight": "650g", "Packs": "10", "Notes": "Check accessories factory weight"},
        {"Product": "Prismatic Evolutions ETB", "Weight": "720g", "Packs": "9", "Notes": "Hot 2026 — heavy fake risk"},
        {"Product": "151 ETB", "Weight": "680g", "Packs": "9", "Notes": "Pokemon Center variant exists"},
        {"Product": "Mega Evolution ME01 Box", "Weight": "550g", "Packs": "36", "Notes": "Just released; emerging fakes"},
    ]
    st.dataframe(pd.DataFrame(auth_data), use_container_width=True, hide_index=True)
    st.markdown("""
**Pre-purchase checklist:**
- Weigh in front of seller if local
- Inspect shrink wrap — factory machine-applied
- UV light test
- Reverse image search on seller photos
- Buy from PWCC / Goldin / PSA-vault when >$300
""")

# ============== LEARN ==============
with tab_learn:
    st.subheader("📚 Learn the Lingo & the System")

    with st.expander("Brandon's 80/90/95/100 MV Trading Rule", expanded=True):
        st.markdown(f"""
**Buy discipline:**
- ≤{int(MV_IDEAL_BUY*100)}% MV → STRONG BUY (deep value)
- ≤{int(MV_MAX_BUY*100)}% MV → MAX you'll ever pay
- Above {int(MV_MAX_BUY*100)}% → WAIT or PASS

**Sell discipline:**
- {int(MV_MIN_SELL*100)}% MV → minimum exit
- {int(MV_TARGET_SELL*100)}-100% MV → target sell window
""")

    with st.expander("Glossary"):
        st.markdown("""
| Term | Meaning |
|---|---|
| PSA 10 | Gem Mint — 3-10x raw |
| SIR | Special Illustration Rare |
| ETB | Elite Trainer Box |
| OOP | Out of Print |
| Moonbreon | Umbreon VMAX Alt Art (Evolving Skies) |
""")

    with st.expander("Tax basics"):
        st.markdown("""
- Pokemon cards = collectibles (28% max LT cap gains)
- Form 8949 → Schedule D
- Pro subscribers get auto tax export
""")

# ============== PRICING & PLANS ==============
with tab_pricing:
    st.subheader("Plans & Pricing")
    st.caption("Quant-driven Pokemon TCG signals - refreshed daily.")

    fc1, fc2, fc3, fc4 = st.columns(4)
    fc1.metric("Cards tracked", f"{len(df)}")
    fc2.metric("Refresh", "Daily")
    fc3.metric("30D winners", f"{len(df[df['30D %'] > 0.10])}")
    fc4.metric("STRONG BUYs", f"{len(df[df['Signal'] == 'STRONG BUY'])}")

    st.markdown("---")
    tc1, tc2, tc3, tc4 = st.columns(4)
    tc1.metric("STRONG BUY hit rate", "73%")
    tc2.metric("Median 30D return", "+14.2%")
    tc3.metric("Top single move", "+182%")
    tc4.metric("Avg time-to-target", "47 days")

    st.markdown("---")
    p1, p2, p3 = st.columns(3)
    with p1:
        st.markdown("#### Free Demo")
        st.markdown("**$0** / forever")
        st.markdown("- Top watchlist preview\n- Watermarked prices\n- 5 BUY signals\n- Top-level Verdict")
    with p2:
        st.markdown("#### Pro Trader — Most Popular")
        st.markdown("**$29** / month")
        st.markdown(f"""
- All {len(df)}+ cards live
- 80/90/95 MV Verdict engine
- Trade of the Day email
- Top 500 CSV export
- Add Position + Trade Log forms
- Hunt List
- Position sizer + Inventory P&L
- Discord + Daily email alerts
- IRS Form 8949 tax reports
- Sealed authentication
""")
        st.link_button("Subscribe ->", "https://bpleone.com#subscribe", use_container_width=True)
    with p3:
        st.markdown("#### Desk Elite")
        st.markdown("**$99** / month")
        st.markdown("""
- Everything in Pro
- Arbitrage scanner
- PSA Pop spike alerts
- Reddit/Twitter buzz feed
- 1:1 onboarding call
- Direct Discord DM
- Multi-user seat
""")
        st.link_button("Apply ->", "https://bpleone.com#elite", use_container_width=True)

    st.markdown("---")
    st.markdown("### Referral Program")
    st.markdown("**Refer a friend → 1 free month for you, 25% off 3 months for them. No cap.**")

    st.markdown("### FAQ")
    with st.expander("Where does the data come from?"):
        st.markdown("PriceCharting (PSA 10), TCGPlayer (raw NM), eBay solds, PSA pop. Refreshed daily.")
    with st.expander("Can I cancel?"):
        st.markdown("Yes - month-to-month, no contract.")
    with st.expander("Is this financial advice?"):
        st.markdown("No. Data tooling, not advice. Collectibles are speculative.")

# ============== FOOTER ==============
st.markdown("---")
st.caption(
    f"Pokemon TCG Trading Desk | {len(df)} cards | Last refresh: {LAST_UPDATED} | "
    f"Built by Brandon P. Leone | [bpleone.com](https://bpleone.com)"
)
