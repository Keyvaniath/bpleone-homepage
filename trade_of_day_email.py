"""
Trade of the Day — Daily Email Subscription System
====================================================
Reads subscribers.csv, finds today's highest-scoring BUY in the workbook,
and emails it to every subscriber whose 'daily' flag is 'yes'.

USAGE
-----
    # Local test (prints email body to stdout, doesn't send):
    DRY_RUN=1 python trade_of_day_email.py

    # Live send (requires env vars):
    GMAIL_USER=... GMAIL_APP_PASSWORD=... python trade_of_day_email.py

GITHUB ACTIONS (workflow runs every morning 7am PT = 14:00 UTC):
---------------------------------------------------------------
    name: Trade of the Day
    on:
      schedule:
        - cron: '0 14 * * *'  # 7am PT daily
      workflow_dispatch:
    jobs:
      send:
        runs-on: ubuntu-latest
        steps:
          - uses: actions/checkout@v4
          - uses: actions/setup-python@v5
            with: { python-version: '3.11' }
          - run: pip install openpyxl pandas
          - run: python trade_of_day_email.py
            env:
              GMAIL_USER: ${{ secrets.GMAIL_USER }}
              GMAIL_APP_PASSWORD: ${{ secrets.GMAIL_APP_PASSWORD }}

SUBSCRIBER CSV FORMAT (subscribers.csv)
---------------------------------------
    email,daily,weekly,alerts,signed_up_at
    user@example.com,yes,yes,no,2026-05-12T10:30:00
"""

import csv
import math
import os
import smtplib
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).parent
WORKBOOK_PATH = ROOT / "Pokemon_Card_Investment_Workbook.xlsx"
SUBSCRIBERS_CSV = ROOT / "subscribers.csv"

GMAIL_USER = os.environ.get("GMAIL_USER", "brandonpleone@gmail.com")
GMAIL_APP_PASSWORD = os.environ.get("GMAIL_APP_PASSWORD", "")
DRY_RUN = os.environ.get("DRY_RUN", "0") == "1"


def compute_score(price, t_buy, d7, d30, pop, tier):
    mom = max(0, min(100, 50 + ((d7 or 0) * 0.6 + (d30 or 0) * 0.4) * 200))
    if t_buy and t_buy > 0:
        val = max(20, min(100, 50 + (t_buy - price) / t_buy * 100))
    else:
        val = 50
    if pop and pop > 0:
        scar = max(0, min(100, 100 - math.log10(pop) * 16))
    else:
        scar = 60
    liq = {"T1": 65, "T2": 85, "T3": 55, "T4": 70, "T5": 45}.get(tier, 50)
    # 55% Momentum / 20% Scarcity / 15% Value / 10% Liquidity
    return mom * 0.55 + scar * 0.20 + val * 0.15 + liq * 0.10


def load_top_card():
    """Return the highest-scoring BUY card from the workbook."""
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    ws = wb["Watchlist"]
    cards = []
    for r in range(5, ws.max_row + 1):
        cid = ws.cell(r, 1).value
        if not cid:
            continue
        price = ws.cell(r, 8).value or 0
        if price <= 0:
            continue
        p7 = ws.cell(r, 10).value or price
        p30 = ws.cell(r, 11).value or price
        t_buy = ws.cell(r, 16).value or 0
        t_sell = ws.cell(r, 17).value or 0
        pop = ws.cell(r, 20).value
        tier = ws.cell(r, 3).value
        d7 = (price - p7) / p7 if p7 else 0
        d30 = (price - p30) / p30 if p30 else 0
        score = compute_score(price, t_buy, d7, d30, pop, tier)

        cards.append({
            "cid": cid,
            "category": ws.cell(r, 2).value,
            "name": ws.cell(r, 4).value,
            "set": ws.cell(r, 5).value,
            "variant": ws.cell(r, 7).value or "",
            "price": price,
            "t_buy": t_buy,
            "t_sell": t_sell,
            "d7": d7,
            "d30": d30,
            "score": round(score, 1),
            "notes": ws.cell(r, 28).value or "",
            "url": ws.cell(r, 27).value or "",
        })
    cards.sort(key=lambda c: c["score"], reverse=True)
    return cards[0] if cards else None


def load_subscribers():
    if not SUBSCRIBERS_CSV.exists():
        return []
    out = []
    with open(SUBSCRIBERS_CSV) as f:
        for row in csv.DictReader(f):
            if (row.get("daily", "").lower() == "yes") and row.get("email"):
                out.append(row["email"].strip())
    return list(set(out))  # dedupe


def build_html_email(card, today):
    """Build the HTML body for the email."""
    mv = card["price"]
    ideal = mv * 0.80
    maxb = mv * 0.90
    target = mv * 0.95

    return f"""<!DOCTYPE html>
<html><body style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; background: #0a0e1a; color: #e8ecf4; padding: 24px;">

<div style="background: #131826; border: 1px solid #f5c842; border-radius: 12px; padding: 32px;">

  <div style="font-family: 'Courier', monospace; color: #f5c842; font-size: 12px; letter-spacing: 0.2em; margin-bottom: 12px;">
    🎯 TRADE OF THE DAY — {today.strftime('%A, %B %d, %Y')}
  </div>

  <h1 style="color: #f5c842; margin: 0 0 8px 0; font-size: 28px;">{card['name']}</h1>
  <div style="color: #8b94a8; font-size: 14px; margin-bottom: 24px;">
    {card['set']} · {card['variant']} · Category: {card['category']}
  </div>

  <div style="background: linear-gradient(135deg, #0a7c2f 0%, #1a9f3a 100%); color: white; padding: 16px; border-radius: 8px; text-align: center; font-size: 22px; font-weight: bold; margin-bottom: 24px;">
    SCORE: {card['score']}/100 — STRONG BUY
  </div>

  <table style="width: 100%; border-collapse: collapse; margin-bottom: 24px;">
    <tr>
      <td style="padding: 12px; border-bottom: 1px solid #232a3e; color: #8b94a8;">Market Value (MV)</td>
      <td style="padding: 12px; border-bottom: 1px solid #232a3e; text-align: right; color: #e8ecf4; font-weight: bold;">${card['price']:,.2f}</td>
    </tr>
    <tr>
      <td style="padding: 12px; border-bottom: 1px solid #232a3e; color: #8b94a8;">Ideal Buy (80% MV)</td>
      <td style="padding: 12px; border-bottom: 1px solid #232a3e; text-align: right; color: #4ade80; font-weight: bold;">${ideal:,.2f}</td>
    </tr>
    <tr>
      <td style="padding: 12px; border-bottom: 1px solid #232a3e; color: #8b94a8;">Max Buy (90% MV)</td>
      <td style="padding: 12px; border-bottom: 1px solid #232a3e; text-align: right; color: #fb923c; font-weight: bold;">${maxb:,.2f}</td>
    </tr>
    <tr>
      <td style="padding: 12px; border-bottom: 1px solid #232a3e; color: #8b94a8;">Target Sell (95% MV)</td>
      <td style="padding: 12px; border-bottom: 1px solid #232a3e; text-align: right; color: #4ade80; font-weight: bold;">${target:,.2f}</td>
    </tr>
    <tr>
      <td style="padding: 12px; color: #8b94a8;">30D Momentum</td>
      <td style="padding: 12px; text-align: right; color: {'#4ade80' if card['d30'] >= 0 else '#f87171'}; font-weight: bold;">{card['d30']:+.1%}</td>
    </tr>
  </table>

  <div style="background: #1a2034; padding: 16px; border-radius: 8px; margin-bottom: 24px; color: #e8ecf4; font-style: italic;">
    📝 {card['notes']}
  </div>

  <div style="text-align: center; margin-bottom: 24px;">
    <a href="https://pokemon.bpleone.com" style="background: #f5c842; color: #0a0e1a; padding: 14px 32px; border-radius: 8px; text-decoration: none; font-weight: bold; display: inline-block;">Open Dashboard →</a>
  </div>

  <div style="text-align: center; margin-bottom: 24px;">
    <a href="{card['url']}" style="color: #f5c842; font-size: 14px; margin: 0 12px;">PriceCharting →</a>
  </div>

  <div style="border-top: 1px solid #232a3e; padding-top: 16px; color: #5a6378; font-size: 11px; text-align: center;">
    Pokemon TCG Trading Desk · bpleone.com<br>
    Not financial advice. Past performance does not guarantee future returns.<br>
    <a href="mailto:brandonpleone@gmail.com?subject=Unsubscribe" style="color: #5a6378;">Unsubscribe</a>
  </div>

</div>

</body></html>"""


def build_plain_email(card, today):
    mv = card["price"]
    return f"""TRADE OF THE DAY — {today.strftime('%A, %B %d, %Y')}

{card['name']}
{card['set']} | {card['variant']}

SCORE: {card['score']}/100 — STRONG BUY

Prices (Brandon's 80/90/95 MV Rule):
  Market Value (MV):    ${mv:,.2f}
  Ideal Buy (80% MV):   ${mv*0.80:,.2f}
  Max Buy (90% MV):     ${mv*0.90:,.2f}
  Target Sell (95% MV): ${mv*0.95:,.2f}

30D Momentum: {card['d30']:+.1%}

Notes: {card['notes']}

Open dashboard: https://pokemon.bpleone.com
PriceCharting:  {card['url']}

---
Pokemon TCG Trading Desk · bpleone.com
Not financial advice. Trade your own conviction.
Reply with "unsubscribe" to opt out.
"""


def send_email(to, subject, html, plain):
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = f"Brandon @ bpleone.com <{GMAIL_USER}>"
    msg["To"] = to
    msg.attach(MIMEText(plain, "plain"))
    msg.attach(MIMEText(html, "html"))

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(GMAIL_USER, GMAIL_APP_PASSWORD)
        server.sendmail(GMAIL_USER, [to], msg.as_string())


def main():
    today = datetime.now()
    card = load_top_card()
    if not card:
        print("No cards found.")
        return

    subject = f"🎯 Trade of the Day — {card['name']} (Score {card['score']})"
    html = build_html_email(card, today)
    plain = build_plain_email(card, today)

    subscribers = load_subscribers()
    print(f"Today's pick: {card['name']} — score {card['score']}")
    print(f"Subscribers to send to: {len(subscribers)}")

    if DRY_RUN:
        print("\n--- PLAIN EMAIL BODY ---\n")
        print(plain)
        print("\n--- (DRY_RUN — no emails sent) ---")
        return

    if not GMAIL_APP_PASSWORD:
        print("ERROR: GMAIL_APP_PASSWORD not set. Refusing to send.")
        return

    sent = 0
    for email in subscribers:
        try:
            send_email(email, subject, html, plain)
            sent += 1
            print(f"  ✓ sent to {email}")
        except Exception as e:
            print(f"  ✗ failed for {email}: {e}")

    print(f"\nDone: {sent}/{len(subscribers)} emails sent.")


if __name__ == "__main__":
    main()
